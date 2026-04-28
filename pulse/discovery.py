from __future__ import annotations

import csv
import json
import xml.etree.ElementTree as ET
import re
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from random import Random
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # pragma: no cover
    load_workbook = None

try:
    import numpy as np
except ModuleNotFoundError:  # pragma: no cover
    np = None

try:
    from PIL import Image
except ModuleNotFoundError:  # pragma: no cover
    Image = None

from .geometry import (
    bbox_union,
    carotid_thickness_from_mask,
    ellipse_circumference,
    load_polygon_json,
    parse_via_polygon,
    polygon_to_bbox,
    safe_literal_eval,
)
from .specs import TaskSpec, TaskType

IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"}
CARDIAC_VIEWS = {"2ch", "4ch", "plax"}
THYROID_BINARY_MAP = {"0": 0, "1": 1}
DOMAIN_ROUTER_DOMAINS = [
    "cardiac",
    "breast",
    "thyroid",
    "fetal",
    "abdominal",
    "liver",
    "kidney",
    "pcos",
    "carotid",
]
DOMAIN_ROUTER_MAX_SAMPLES_PER_DOMAIN = 4200


def _list_images(path: Path) -> List[Path]:
    if not path.exists():
        return []
    return [item for item in sorted(path.iterdir()) if item.suffix.lower() in IMAGE_SUFFIXES]


def _stratified_split(
    samples: List[Dict],
    seed: int,
    label_key: str = "label",
    val_ratio: float = 0.15,
    test_ratio: float = 0.15,
) -> Tuple[List[Dict], List[Dict], List[Dict]]:
    rng = Random(seed)
    grouped: Dict[int, List[Dict]] = defaultdict(list)
    for sample in samples:
        grouped[int(sample[label_key])].append(sample)

    train: List[Dict] = []
    val: List[Dict] = []
    test: List[Dict] = []
    for label_samples in grouped.values():
        rng.shuffle(label_samples)
        total = len(label_samples)
        test_count = max(1, int(total * test_ratio)) if total >= 8 else max(0, int(total * test_ratio))
        val_count = max(1, int(total * val_ratio)) if total >= 8 else max(0, int(total * val_ratio))
        test.extend(label_samples[:test_count])
        val.extend(label_samples[test_count:test_count + val_count])
        train.extend(label_samples[test_count + val_count:])
    return train, val, test


def _group_split(
    samples: List[Dict],
    seed: int,
    group_key: str = "patient_id",
    val_ratio: float = 0.15,
    test_ratio: float = 0.15,
) -> Tuple[List[Dict], List[Dict], List[Dict]]:
    rng = Random(seed)
    buckets: Dict[str, List[Dict]] = defaultdict(list)
    for sample in samples:
        buckets[str(sample[group_key])].append(sample)
    group_ids = list(buckets)
    rng.shuffle(group_ids)
    total = len(group_ids)
    test_cut = int(total * test_ratio)
    val_cut = int(total * val_ratio)
    test_ids = set(group_ids[:test_cut])
    val_ids = set(group_ids[test_cut:test_cut + val_cut])
    train_ids = set(group_ids[test_cut + val_cut:])
    train, val, test = [], [], []
    for group_id in train_ids:
        train.extend(buckets[group_id])
    for group_id in val_ids:
        val.extend(buckets[group_id])
    for group_id in test_ids:
        test.extend(buckets[group_id])
    return train, val, test


def _group_stratified_split(
    samples: List[Dict],
    seed: int,
    label_key: str = "label",
    group_key: str = "patient_id",
    val_ratio: float = 0.15,
    test_ratio: float = 0.15,
) -> Tuple[List[Dict], List[Dict], List[Dict]]:
    rng = Random(seed)
    group_buckets: Dict[str, List[Dict]] = defaultdict(list)
    for sample in samples:
        group_buckets[str(sample[group_key])].append(sample)

    label_to_groups: Dict[int, List[str]] = defaultdict(list)
    for group_id, group_samples in group_buckets.items():
        labels = [int(sample[label_key]) for sample in group_samples if label_key in sample]
        if not labels:
            continue
        majority_label = Counter(labels).most_common(1)[0][0]
        label_to_groups[majority_label].append(group_id)

    train_ids = set()
    val_ids = set()
    test_ids = set()
    for group_ids in label_to_groups.values():
        rng.shuffle(group_ids)
        total = len(group_ids)
        test_count = max(1, int(total * test_ratio)) if total >= 8 else max(0, int(total * test_ratio))
        val_count = max(1, int(total * val_ratio)) if total >= 8 else max(0, int(total * val_ratio))
        test_ids.update(group_ids[:test_count])
        val_ids.update(group_ids[test_count:test_count + val_count])
        train_ids.update(group_ids[test_count + val_count:])

    train: List[Dict] = []
    val: List[Dict] = []
    test: List[Dict] = []
    for group_id in train_ids:
        train.extend(group_buckets[group_id])
    for group_id in val_ids:
        val.extend(group_buckets[group_id])
    for group_id in test_ids:
        test.extend(group_buckets[group_id])
    return train, val, test


def _split_train_and_val(
    samples: List[Dict],
    seed: int,
    label_key: Optional[str] = "label",
    group_key: Optional[str] = None,
    val_ratio: float = 0.15,
) -> Tuple[List[Dict], List[Dict]]:
    if group_key:
        train, val, _ = _group_split(samples, seed=seed, group_key=group_key, val_ratio=val_ratio, test_ratio=0.0)
        return train, val
    if label_key:
        train, val, _ = _stratified_split(samples, seed=seed, label_key=label_key, val_ratio=val_ratio, test_ratio=0.0)
        return train, val
    rng = Random(seed)
    samples = list(samples)
    rng.shuffle(samples)
    cut = int(len(samples) * val_ratio)
    return samples[cut:], samples[:cut]


def _limit_samples_by_group(
    samples: List[Dict],
    seed: int,
    max_samples: int,
    group_key: str = "patient_id",
) -> List[Dict]:
    if max_samples <= 0 or len(samples) <= max_samples:
        return list(samples)

    rng = Random(seed)
    grouped: Dict[str, List[Dict]] = defaultdict(list)
    for sample in samples:
        group_id = str(sample.get(group_key) or Path(str(sample.get("image", ""))).stem)
        grouped[group_id].append(sample)

    group_ids = list(grouped)
    rng.shuffle(group_ids)
    limited: List[Dict] = []
    for group_id in group_ids:
        group_samples = grouped[group_id]
        if not limited or len(limited) + len(group_samples) <= max_samples:
            limited.extend(group_samples)
        if len(limited) >= max_samples:
            break

    return limited[:max_samples]


def _read_csv_rows(path: Path, delimiter: str = ",") -> List[Dict[str, str]]:
    with open(path, newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle, delimiter=delimiter))


def _read_xlsx_rows(path: Path) -> List[Dict[str, object]]:
    if load_workbook is None:
        return _read_xlsx_rows_fallback(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    data = []
    for row in rows[1:]:
        data.append({header: value for header, value in zip(headers, row)})
    return data


def _xlsx_column_index(cell_ref: str) -> int:
    match = re.match(r"([A-Z]+)", cell_ref.upper())
    if not match:
        return 0
    value = 0
    for char in match.group(1):
        value = (value * 26) + (ord(char) - ord("A") + 1)
    return value - 1


def _read_xlsx_rows_fallback(path: Path) -> List[Dict[str, object]]:
    if not path.exists():
        return []
    namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    rel_namespace = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}
    with zipfile.ZipFile(path, "r") as archive:
        shared_strings: List[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for si in root.findall("main:si", namespace):
                text_parts = [node.text or "" for node in si.findall(".//main:t", namespace)]
                shared_strings.append("".join(text_parts))

        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        sheets = workbook.find("main:sheets", namespace)
        if sheets is None:
            return []
        first_sheet = sheets.find("main:sheet", namespace)
        if first_sheet is None:
            return []
        rel_id = first_sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        worksheet_path = "xl/worksheets/sheet1.xml"
        if rel_id and "xl/_rels/workbook.xml.rels" in archive.namelist():
            rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
            for rel in rels.findall("rel:Relationship", rel_namespace):
                if rel.attrib.get("Id") == rel_id:
                    target = rel.attrib.get("Target", "worksheets/sheet1.xml")
                    worksheet_path = f"xl/{target.lstrip('/')}"
                    break

        sheet_root = ET.fromstring(archive.read(worksheet_path))
        sheet_data = sheet_root.find("main:sheetData", namespace)
        if sheet_data is None:
            return []

        rows: List[List[object]] = []
        for row in sheet_data.findall("main:row", namespace):
            cells: Dict[int, object] = {}
            for cell in row.findall("main:c", namespace):
                cell_ref = cell.attrib.get("r", "A1")
                column_index = _xlsx_column_index(cell_ref)
                cell_type = cell.attrib.get("t")
                value_node = cell.find("main:v", namespace)
                inline_node = cell.find("main:is", namespace)
                value: object = ""
                if inline_node is not None:
                    text_parts = [node.text or "" for node in inline_node.findall(".//main:t", namespace)]
                    value = "".join(text_parts)
                elif value_node is not None and value_node.text is not None:
                    raw = value_node.text
                    if cell_type == "s":
                        shared_index = int(raw)
                        value = shared_strings[shared_index] if 0 <= shared_index < len(shared_strings) else raw
                    else:
                        value = raw
                cells[column_index] = value
            if not cells:
                continue
            width = max(cells) + 1
            rows.append([cells.get(index, "") for index in range(width)])

    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    data: List[Dict[str, object]] = []
    for row in rows[1:]:
        data.append({header: value for header, value in zip(headers, row)})
    return data


def _busbra_patient_id(raw_id: str) -> str:
    return raw_id.rsplit("-", 1)[0]


def _kidney_patient_id(filename: str) -> str:
    stem = Path(filename).stem
    return stem.split("_", 1)[0]


def _pcos_patient_id(image_name: str) -> str:
    stem = Path(image_name).stem
    if stem.startswith(("normal_", "pco_")):
        return stem
    match = re.match(r"^(?P<prefix>\d+)image", stem)
    if match:
        return match.group("prefix")
    return stem


def _group_stratified_val_test_split(
    samples: List[Dict],
    seed: int,
    label_key: str = "label",
    group_key: str = "patient_id",
    val_ratio: float = 0.5,
) -> Tuple[List[Dict], List[Dict]]:
    train, val, test = _group_stratified_split(
        samples,
        seed=seed,
        label_key=label_key,
        group_key=group_key,
        val_ratio=val_ratio,
        test_ratio=max(0.0, 1.0 - val_ratio),
    )
    val_samples = [*train, *val]
    return val_samples, test


def _build_unavailable_task(task_id: str, domain: str, title: str, task_type: TaskType, dataset_names: Sequence[str], reason: str) -> TaskSpec:
    return TaskSpec(
        task_id=task_id,
        domain=domain,
        title=title,
        task_type=task_type,
        dataset_names=list(dataset_names),
        status="unavailable",
        skip_reason=reason,
    )


def _router_image_path(sample: Dict) -> Optional[str]:
    image = sample.get("image")
    if image:
        return str(image)
    modalities = sample.get("modalities")
    if isinstance(modalities, dict):
        if modalities.get("bmode"):
            return str(modalities["bmode"])
        for key in sorted(modalities):
            if modalities.get(key):
                return str(modalities[key])
    return None


def _router_label_name(task: TaskSpec, sample: Dict) -> str:
    label_name = sample.get("label_name")
    if isinstance(label_name, str) and label_name.strip():
        return label_name.strip().lower()
    label = sample.get("label")
    if isinstance(label, int) and 0 <= label < len(task.labels):
        return str(task.labels[label]).strip().lower()
    return ""


def _router_sample_domain(task: TaskSpec, sample: Dict) -> Optional[str]:
    if task.task_id == "cardiac/butterfly_cardiac_screening":
        return None

    label_name = _router_label_name(task, sample)
    if task.task_id == "cardiac/butterfly_view_classification":
        butterfly_map = {
            "2ch": "cardiac",
            "4ch": "cardiac",
            "plax": "cardiac",
            "thyroid": "thyroid",
            "carotid": "carotid",
            "bladder": "abdominal",
            "ivc": "abdominal",
            "morisons_pouch": "abdominal",
        }
        return butterfly_map.get(label_name)

    if task.task_id == "abdominal/organ_classification":
        abdominal_map = {
            "aa": "abdominal",
            "gb": "abdominal",
            "hepatic": "liver",
            "kidney": "kidney",
            "liver": "liver",
            "pancreas": "abdominal",
            "portal": "abdominal",
            "spleen": "abdominal",
            "ub+prostrate+uterous+cervix": "abdominal",
        }
        return abdominal_map.get(label_name)

    if task.task_id == "abdominal/anomaly_detection":
        return None

    return task.domain


def _aggregate_router_domain_samples(tasks: Sequence[TaskSpec], domain: str, seed: int) -> List[Dict]:
    aggregated: List[Dict] = []
    seen: set[Tuple[str, str]] = set()
    for task in tasks:
        if not task.is_ready or task.domain == "system":
            continue
        source_samples = [*task.train_samples, *task.val_samples, *task.test_samples]
        for sample in source_samples:
            sample_domain = _router_sample_domain(task, sample)
            if sample_domain != domain:
                continue
            image_path = _router_image_path(sample)
            if not image_path:
                continue
            patient_id = f"{domain}_{sample.get('patient_id', Path(image_path).stem)}"
            dedupe_key = (str(Path(image_path).resolve()), patient_id)
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            aggregated.append(
                {
                    "image": str(Path(image_path).resolve()),
                    "patient_id": patient_id,
                    "source_task": task.task_id,
                }
            )
    return _limit_samples_by_group(
        aggregated,
        seed=seed,
        max_samples=DOMAIN_ROUTER_MAX_SAMPLES_PER_DOMAIN,
        group_key="patient_id",
    )


def discover_butterfly_tasks(data_root: Path, seed: int) -> List[TaskSpec]:
    root = data_root / "Butterfly"
    if not root.exists():
        return [
            _build_unavailable_task(
                "cardiac/butterfly_view_classification",
                "cardiac",
                "Butterfly view classification",
                TaskType.CLASSIFICATION,
                ["Butterfly GrandHack 2018"],
                "Butterfly dataset folder was not found.",
            )
        ]

    train_pool: List[Dict] = []
    test_samples: List[Dict] = []
    labels: List[str] = []
    for patient_dir in sorted(item for item in root.iterdir() if item.is_dir() and item.name != "Test"):
        for view_dir in sorted(item for item in patient_dir.iterdir() if item.is_dir()):
            labels.append(view_dir.name)
            for image_path in _list_images(view_dir):
                train_pool.append(
                    {
                        "image": str(image_path),
                        "label_name": view_dir.name,
                        "patient_id": patient_dir.name,
                    }
                )

    test_root = root / "Test"
    if test_root.exists():
        for patient_dir in sorted(item for item in test_root.iterdir() if item.is_dir()):
            for view_dir in sorted(item for item in patient_dir.iterdir() if item.is_dir()):
                labels.append(view_dir.name)
                for image_path in _list_images(view_dir):
                    test_samples.append(
                        {
                            "image": str(image_path),
                            "label_name": view_dir.name,
                            "patient_id": f"test_{patient_dir.name}",
                        }
                    )

    label_names = sorted(set(labels))
    label_map = {name: index for index, name in enumerate(label_names)}
    for sample in train_pool:
        sample["label"] = label_map[sample["label_name"]]
    for sample in test_samples:
        sample["label"] = label_map[sample["label_name"]]

    train_samples, val_samples = _split_train_and_val(train_pool, seed=seed, group_key="patient_id")
    view_task = TaskSpec(
        task_id="cardiac/butterfly_view_classification",
        domain="cardiac",
        title="Butterfly 9-view classification",
        task_type=TaskType.CLASSIFICATION,
        dataset_names=["Butterfly GrandHack 2018"],
        labels=label_names,
        description="9-class multi-POCUS view classification over Butterfly clips collapsed to frames.",
        train_samples=train_samples,
        val_samples=val_samples,
        test_samples=test_samples,
    )

    screening_train = [{**sample, "label": int(sample["label_name"] in CARDIAC_VIEWS)} for sample in train_samples]
    screening_val = [{**sample, "label": int(sample["label_name"] in CARDIAC_VIEWS)} for sample in val_samples]
    screening_test = [{**sample, "label": int(sample["label_name"] in CARDIAC_VIEWS)} for sample in test_samples]
    screening_task = TaskSpec(
        task_id="cardiac/butterfly_cardiac_screening",
        domain="cardiac",
        title="Butterfly cardiac screening proxy",
        task_type=TaskType.CLASSIFICATION,
        dataset_names=["Butterfly GrandHack 2018"],
        labels=["non_cardiac", "cardiac"],
        description="Binary screening proxy that groups 2CH, 4CH, and PLAX as cardiac-positive views.",
        train_samples=screening_train,
        val_samples=screening_val,
        test_samples=screening_test,
        notes=["This is a lightweight cardiac-screening proxy derived from view labels."],
    )
    return [view_task, screening_task]


def discover_breast_tasks(data_root: Path, seed: int) -> List[TaskSpec]:
    label_map = {"normal": 0, "benign": 1, "malignant": 2}
    classification_samples: List[Dict] = []
    segmentation_samples: List[Dict] = []

    busi_root = data_root / "Breast_BUSI_with_GT"
    if busi_root.exists():
        for label_name, label_value in label_map.items():
            class_dir = busi_root / label_name
            for image_path in _list_images(class_dir):
                if "_mask" in image_path.stem:
                    continue
                mask_paths = sorted(class_dir.glob(f"{image_path.stem}_mask*.png"))
                classification_samples.append(
                    {
                        "image": str(image_path),
                        "label": label_value,
                        "patient_id": f"busi_{image_path.stem}",
                    }
                )
                if mask_paths:
                    segmentation_samples.append(
                        {
                            "image": str(image_path),
                            "mask_paths": [str(path) for path in mask_paths],
                            "patient_id": f"busi_{image_path.stem}",
                        }
                    )

    busbra_images = data_root / "Breast - BUSBRA" / "Images"
    busbra_masks = data_root / "Breast - BUSBRA" / "Masks"
    busbra_split_csv = data_root / "Breast - BUSBRA" / "5-fold-cv.csv"
    if busbra_images.exists() and busbra_split_csv.exists():
        for row in _read_csv_rows(busbra_split_csv):
            pathology = row["Pathology"].strip().lower()
            image_path = busbra_images / f"{row['ID']}.png"
            if image_path.exists():
                classification_samples.append(
                    {
                        "image": str(image_path),
                        "label": label_map[pathology],
                        "patient_id": f"busbra_{_busbra_patient_id(row['ID'])}",
                    }
                )
                mask_path = busbra_masks / f"mask_{row['ID'].replace('bus_', '')}.png"
                if not mask_path.exists():
                    mask_path = busbra_masks / f"mask_{row['ID'].split('bus_')[-1]}.png"
                if mask_path.exists():
                    segmentation_samples.append(
                        {
                            "image": str(image_path),
                            "mask_path": str(mask_path),
                            "patient_id": f"busbra_{_busbra_patient_id(row['ID'])}",
                        }
                    )

    bus_uc_cls_root = data_root / "Breast_BUS_UC" / "BUS_UC_classification" / "BUS_UC_classification"
    if bus_uc_cls_root.exists():
        for label_name in ("Benign", "Malignant"):
            label_value = label_map[label_name.lower()]
            for image_path in _list_images(bus_uc_cls_root / label_name):
                classification_samples.append(
                    {
                        "image": str(image_path),
                        "label": label_value,
                        "patient_id": f"busuccls_{label_name}_{image_path.stem}",
                    }
                )

    bus_uc_seg_root = data_root / "Breast_BUS_UC" / "BUS_UC" / "BUS_UC" / "BUS_UC"
    if bus_uc_seg_root.exists():
        for label_name in ("Benign", "Malignant"):
            image_dir = bus_uc_seg_root / label_name / "images"
            mask_dir = bus_uc_seg_root / label_name / "masks"
            for image_path in _list_images(image_dir):
                mask_path = mask_dir / image_path.name
                if mask_path.exists():
                    segmentation_samples.append(
                        {
                            "image": str(image_path),
                            "mask_path": str(mask_path),
                            "patient_id": f"busucseg_{label_name}_{image_path.stem}",
                        }
                    )

    clf_train, clf_val, clf_test = _group_stratified_split(
        classification_samples,
        seed=seed,
        label_key="label",
        group_key="patient_id",
    )
    seg_train, seg_val, seg_test = _group_split(segmentation_samples, seed=seed, group_key="patient_id")
    return [
        TaskSpec(
            task_id="breast/lesion_classification",
            domain="breast",
            title="Breast lesion classification",
            task_type=TaskType.CLASSIFICATION,
            dataset_names=["BUSI", "BUSBRA", "BUS-UC"],
            labels=["normal", "benign", "malignant"],
            description="Merged 3-class lesion classification over BUSI, BUSBRA, and BUS-UC.",
            train_samples=clf_train,
            val_samples=clf_val,
            test_samples=clf_test,
        ),
        TaskSpec(
            task_id="breast/lesion_segmentation",
            domain="breast",
            title="Breast lesion segmentation",
            task_type=TaskType.SEGMENTATION,
            dataset_names=["BUSI", "BUSBRA", "BUS-UC"],
            labels=["lesion"],
            description="Binary lesion segmentation over BUSI, BUSBRA, and BUS-UC masks.",
            train_samples=seg_train,
            val_samples=seg_val,
            test_samples=seg_test,
            extras={"segmentation_mode": "binary"},
        ),
    ]


def _parse_voc_boxes(xml_path: Path) -> Tuple[Optional[List[float]], Tuple[int, int]]:
    tree = ET.parse(xml_path)
    root = tree.getroot()
    size_node = root.find("size")
    width = int(size_node.findtext("width", "1")) if size_node is not None else 1
    height = int(size_node.findtext("height", "1")) if size_node is not None else 1
    boxes = []
    for obj in root.findall("object"):
        bbox = obj.find("bndbox")
        if bbox is None:
            continue
        boxes.append(
            [
                float(bbox.findtext("xmin", "0")),
                float(bbox.findtext("ymin", "0")),
                float(bbox.findtext("xmax", "0")),
                float(bbox.findtext("ymax", "0")),
            ]
        )
    return (bbox_union(boxes) if boxes else None, (width, height))


def discover_thyroid_tasks(data_root: Path, seed: int) -> List[TaskSpec]:
    classification_samples: List[Dict] = []
    multimodal_samples: List[Dict] = []

    kaggle_root = data_root / "Thyroid - Kaggle"
    if kaggle_root.exists():
        for label_name, label_value in {"benign": 0, "malignant": 1}.items():
            for image_path in _list_images(kaggle_root / label_name):
                classification_samples.append(
                    {
                        "image": str(image_path),
                        "label": label_value,
                        "patient_id": f"kaggle_{image_path.stem}",
                    }
                )

    multimodal_root = data_root / "Thyroid_dataset"
    if multimodal_root.exists():
        modality_maps: Dict[str, Dict[str, Dict[str, Path]]] = defaultdict(lambda: defaultdict(dict))

        def case_key(path: Path) -> str:
            parts = path.stem.rsplit("-", 1)
            return parts[0] if len(parts) == 2 else path.stem

        for modality in ("2d", "cdfi", "elastic"):
            modality_dir = multimodal_root / modality
            if not modality_dir.exists():
                continue
            for label_dir in sorted(item for item in modality_dir.iterdir() if item.is_dir()):
                label_name = label_dir.name
                for image_path in _list_images(label_dir):
                    modality_maps[label_name][case_key(image_path)][modality] = image_path
        for label_name, sample_map in modality_maps.items():
            for stem, modalities in sample_map.items():
                if {"2d", "cdfi", "elastic"} <= set(modalities):
                    multimodal_samples.append(
                        {
                            "modalities": {
                                "bmode": str(modalities["2d"]),
                                "cdfi": str(modalities["cdfi"]),
                                "elastic": str(modalities["elastic"]),
                            },
                            "label": THYROID_BINARY_MAP.get(label_name, 0),
                            "patient_id": f"thyroidmm_{stem}",
                        }
                    )
                    classification_samples.append(
                        {
                            "image": str(modalities["2d"]),
                            "label": THYROID_BINARY_MAP.get(label_name, 0),
                            "patient_id": f"thyroid2d_{stem}",
                        }
                    )

    cls_train, cls_val, cls_test = _group_stratified_split(
        classification_samples,
        seed=seed,
        label_key="label",
        group_key="patient_id",
    )
    mm_train, mm_val, mm_test = _group_stratified_split(
        multimodal_samples,
        seed=seed,
        label_key="label",
        group_key="patient_id",
    )

    detection_root = data_root / "Thyroid - TN5000" / "TN5000_forReview"
    detection_task: TaskSpec
    if detection_root.exists():
        image_dir = detection_root / "JPEGImages"
        ann_dir = detection_root / "Annotations"
        train_list = detection_root / "ImageSets" / "Main" / "trainval.txt"
        test_list = detection_root / "ImageSets" / "Main" / "test.txt"

        def build_detection_samples(names: Iterable[str]) -> List[Dict]:
            samples: List[Dict] = []
            for name in names:
                image_path = image_dir / f"{name}.jpg"
                xml_path = ann_dir / f"{name}.xml"
                if not image_path.exists() or not xml_path.exists():
                    continue
                bbox, _ = _parse_voc_boxes(xml_path)
                if bbox is None:
                    continue
                samples.append(
                    {
                        "image": str(image_path),
                        "bbox": bbox,
                        "label": 1,
                        "patient_id": name,
                    }
                )
            return samples

        if train_list.exists() and test_list.exists():
            train_names = [line.strip() for line in train_list.read_text(encoding="utf-8").splitlines() if line.strip()]
            test_names = [line.strip() for line in test_list.read_text(encoding="utf-8").splitlines() if line.strip()]
            train_pool = build_detection_samples(train_names)
            det_train, det_val = _split_train_and_val(train_pool, seed=seed, group_key="patient_id")
            det_test = build_detection_samples(test_names)
        else:
            all_names = [path.stem for path in ann_dir.glob("*.xml")]
            all_samples = build_detection_samples(all_names)
            det_train, det_val, det_test = _group_split(all_samples, seed=seed, group_key="patient_id")

        detection_task = TaskSpec(
            task_id="thyroid/nodule_detection",
            domain="thyroid",
            title="Thyroid single-nodule detection",
            task_type=TaskType.DETECTION,
            dataset_names=["TN5000_forReview"],
            labels=["nodule"],
            description="Single-box nodule detection from TN5000 VOC annotations.",
            train_samples=det_train,
            val_samples=det_val,
            test_samples=det_test,
            extras={"single_box": True},
        )
    else:
        detection_task = _build_unavailable_task(
            "thyroid/nodule_detection",
            "thyroid",
            "Thyroid single-nodule detection",
            TaskType.DETECTION,
            ["TN5000_forReview"],
            "TN5000_forReview annotations were not found.",
        )

    if mm_train or mm_val or mm_test:
        multimodal_task = TaskSpec(
            task_id="thyroid/multimodal_fusion_classification",
            domain="thyroid",
            title="Thyroid multimodal fusion classification",
            task_type=TaskType.MULTIMODAL,
            dataset_names=["Thyroid_dataset"],
            labels=["benign", "malignant"],
            description="Three-branch fusion over B-mode, CDFI, and elastography inputs.",
            train_samples=mm_train,
            val_samples=mm_val,
            test_samples=mm_test,
            extras={"modalities": ["bmode", "cdfi", "elastic"]},
            notes=["The local Thyroid_dataset folder is assumed to map class 0->benign and class 1->malignant."],
        )
    else:
        multimodal_task = _build_unavailable_task(
            "thyroid/multimodal_fusion_classification",
            "thyroid",
            "Thyroid multimodal fusion classification",
            TaskType.MULTIMODAL,
            ["Thyroid_dataset"],
            "No aligned 2D/CDFI/elastography triplets were discovered.",
        )

    return [
        TaskSpec(
            task_id="thyroid/benign_malignant_classification",
            domain="thyroid",
            title="Thyroid benign/malignant classification",
            task_type=TaskType.CLASSIFICATION,
            dataset_names=["Thyroid Kaggle", "Thyroid_dataset"],
            labels=["benign", "malignant"],
            description="Binary thyroid classification from Kaggle 2D US and the multimodal dataset 2D branch.",
            train_samples=cls_train,
            val_samples=cls_val,
            test_samples=cls_test,
            notes=["The local Thyroid_dataset folder is assumed to map class 0->benign and class 1->malignant."],
        ),
        detection_task,
        multimodal_task,
    ]


def discover_fetal_tasks(data_root: Path, seed: int) -> List[TaskSpec]:
    fetal_plane_root = data_root / "FETAL_PLANES_ZENODO"
    plane_task: TaskSpec
    if fetal_plane_root.exists():
        csv_path = fetal_plane_root / "FETAL_PLANES_DB_data.csv"
        rows = _read_csv_rows(csv_path, delimiter=";")
        train_pool: List[Dict] = []
        test_samples: List[Dict] = []
        label_names = sorted({row["Plane"].strip() for row in rows if row.get("Plane")})
        label_map = {name: index for index, name in enumerate(label_names)}
        for row in rows:
            image_name = row["Image_name"].strip()
            image_path = fetal_plane_root / "Images" / f"{image_name}.png"
            if not image_path.exists():
                continue
            sample = {
                "image": str(image_path),
                "label": label_map[row["Plane"].strip()],
                "patient_id": row["Patient_num"].strip(),
            }
            is_train = row.get("Train ", row.get("Train", "1")).strip() == "1"
            if is_train:
                train_pool.append(sample)
            else:
                test_samples.append(sample)
        train_samples, val_samples = _split_train_and_val(train_pool, seed=seed, group_key="patient_id")
        plane_task = TaskSpec(
            task_id="fetal/plane_classification",
            domain="fetal",
            title="Fetal plane classification",
            task_type=TaskType.CLASSIFICATION,
            dataset_names=["FETAL_PLANES"],
            labels=label_names,
            description="6-class fetal plane classification from the FETAL_PLANES benchmark.",
            train_samples=train_samples,
            val_samples=val_samples,
            test_samples=test_samples,
        )
    else:
        plane_task = _build_unavailable_task(
            "fetal/plane_classification",
            "fetal",
            "Fetal plane classification",
            TaskType.CLASSIFICATION,
            ["FETAL_PLANES"],
            "FETAL_PLANES_ZENODO folder was not found.",
        )

    hc_root = data_root / "Fetal_HC"
    hc_task: TaskSpec
    if hc_root.exists():
        rows = _read_csv_rows(hc_root / "training_set_pixel_size_and_HC.csv")
        hc_samples = []
        for row in rows:
            image_path = hc_root / "training_set" / row["filename"].strip()
            if image_path.exists():
                hc_samples.append(
                    {
                        "image": str(image_path),
                        "target": float(row["head circumference (mm)"]),
                        "pixel_size": float(row["pixel size(mm)"]),
                        "patient_id": image_path.stem,
                    }
                )
        hc_train, hc_val, hc_test = _group_split(hc_samples, seed=seed, group_key="patient_id")
        hc_task = TaskSpec(
            task_id="fetal/hc_measurement",
            domain="fetal",
            title="Fetal head circumference measurement",
            task_type=TaskType.MEASUREMENT,
            dataset_names=["Fetal_HC"],
            description="Regression over fetal head circumference targets in millimeters.",
            train_samples=hc_train,
            val_samples=hc_val,
            test_samples=hc_test,
            extras={"target_name": "head_circumference_mm"},
            notes=["The official unlabeled Fetal_HC test split is left untouched; this task uses internal labeled splits."],
        )
    else:
        hc_task = _build_unavailable_task(
            "fetal/hc_measurement",
            "fetal",
            "Fetal head circumference measurement",
            TaskType.MEASUREMENT,
            ["Fetal_HC"],
            "Fetal_HC folder was not found.",
        )

    return [plane_task, hc_task]


def discover_abdominal_tasks(data_root: Path, seed: int) -> List[TaskSpec]:
    organ_root_1 = data_root / "Organ_Anomaly" / "organ_classification_1"
    organ_root_2 = data_root / "Organ_Anomaly" / "organ_classification_2"
    anomaly_root = data_root / "Organ_Anomaly" / "organ_classification+anomaly_detection"

    class_names = sorted(
        {
            path.name
            for root in (organ_root_1, organ_root_2)
            if root.exists()
            for path in root.iterdir()
            if path.is_dir()
        }
    )
    label_map = {name: index for index, name in enumerate(class_names)}
    classification_samples: List[Dict] = []
    for dataset_name, folder in [("organ1", organ_root_1), ("organ2", organ_root_2)]:
        if not folder.exists():
            continue
        for class_dir in sorted(item for item in folder.iterdir() if item.is_dir()):
            for image_path in _list_images(class_dir):
                classification_samples.append(
                    {
                        "image": str(image_path),
                        "label": label_map[class_dir.name],
                        "patient_id": f"{dataset_name}_{class_dir.name}_{image_path.stem}",
                    }
                )

    anomaly_samples: List[Dict] = []
    if anomaly_root.exists():
        for anomaly_dir in sorted(item for item in anomaly_root.iterdir() if item.is_dir()):
            label = 1 if anomaly_dir.name.lower() == "abnormal" else 0
            for organ_dir in sorted(item for item in anomaly_dir.iterdir() if item.is_dir()):
                for image_path in _list_images(organ_dir):
                    anomaly_samples.append(
                        {
                            "image": str(image_path),
                            "label": label,
                            "patient_id": f"{anomaly_dir.name}_{organ_dir.name}_{image_path.stem}",
                        }
                    )

    clf_train, clf_val, clf_test = _group_stratified_split(
        classification_samples,
        seed=seed,
        label_key="label",
        group_key="patient_id",
    )
    an_train, an_val, an_test = _group_stratified_split(
        anomaly_samples,
        seed=seed,
        label_key="label",
        group_key="patient_id",
    )
    return [
        TaskSpec(
            task_id="abdominal/organ_classification",
            domain="abdominal",
            title="Abdominal organ classification",
            task_type=TaskType.CLASSIFICATION,
            dataset_names=["Organ Anomaly"],
            labels=class_names,
            description="10-class abdominal organ classification over the Organ Anomaly subsets.",
            train_samples=clf_train,
            val_samples=clf_val,
            test_samples=clf_test,
        ),
        TaskSpec(
            task_id="abdominal/anomaly_detection",
            domain="abdominal",
            title="Abdominal anomaly detection",
            task_type=TaskType.CLASSIFICATION,
            dataset_names=["Organ Anomaly"],
            labels=["normal", "abnormal"],
            description="Binary anomaly detection over the mixed normal/abnormal abdominal organ subset.",
            train_samples=an_train,
            val_samples=an_val,
            test_samples=an_test,
        ),
    ]


def discover_liver_tasks(data_root: Path, seed: int) -> List[TaskSpec]:
    root = data_root / "Liver"
    class_names = ["Benign", "Malignant", "Normal"]
    label_map = {name: index for index, name in enumerate(class_names)}
    classification_samples: List[Dict] = []
    segmentation_samples: List[Dict] = []
    for class_name in class_names:
        image_dir = root / class_name / "image"
        liver_json_dir = root / class_name / "segmentation" / "liver"
        outline_json_dir = root / class_name / "segmentation" / "outline"
        if not image_dir.exists():
            continue
        for image_path in _list_images(image_dir):
            classification_samples.append(
                {
                    "image": str(image_path),
                    "label": label_map[class_name],
                    "patient_id": f"liver_{class_name}_{image_path.stem}",
                }
            )
            polygons_by_class = {}
            liver_json = liver_json_dir / f"{image_path.stem}.json"
            outline_json = outline_json_dir / f"{image_path.stem}.json"
            if liver_json.exists():
                polygons_by_class[1] = load_polygon_json(str(liver_json))
            if outline_json.exists():
                polygons_by_class[2] = load_polygon_json(str(outline_json))
            if polygons_by_class:
                segmentation_samples.append(
                    {
                        "image": str(image_path),
                        "mask_polygons": polygons_by_class,
                        "patient_id": f"liverseg_{class_name}_{image_path.stem}",
                    }
                )
    clf_train, clf_val, clf_test = _group_stratified_split(
        classification_samples,
        seed=seed,
        label_key="label",
        group_key="patient_id",
    )
    seg_train, seg_val, seg_test = _group_split(segmentation_samples, seed=seed, group_key="patient_id")
    return [
        TaskSpec(
            task_id="liver/pathology_classification",
            domain="liver",
            title="Liver pathology classification",
            task_type=TaskType.CLASSIFICATION,
            dataset_names=["Liver US"],
            labels=[name.lower() for name in class_names],
            description="3-class liver pathology classification.",
            train_samples=clf_train,
            val_samples=clf_val,
            test_samples=clf_test,
        ),
        TaskSpec(
            task_id="liver/segmentation",
            domain="liver",
            title="Liver and lesion segmentation",
            task_type=TaskType.SEGMENTATION,
            dataset_names=["Liver US"],
            labels=["liver", "lesion"],
            description="Multi-class segmentation with liver region and lesion outline channels.",
            train_samples=seg_train,
            val_samples=seg_val,
            test_samples=seg_test,
            extras={
                "segmentation_mode": "multiclass",
                "mask_class_names": ["background", "liver", "lesion"],
            },
        ),
    ]


def discover_kidney_tasks(data_root: Path, seed: int) -> List[TaskSpec]:
    kidney_root = data_root / "kidneyUS-main"
    segmentation_task: TaskSpec
    if kidney_root.exists():
        label_csv = kidney_root / "labels" / "reviewed_labels_1.csv"
        image_dir = kidney_root / "kidneyUS_images_25_june_2025"
        anatomy_map = {
            "capsule": 1,
            "central echo complex": 2,
            "medulla": 3,
            "cortex": 4,
        }
        polygon_store: Dict[str, Dict[int, List[List[Tuple[float, float]]]]] = defaultdict(lambda: defaultdict(list))
        for row in _read_csv_rows(label_csv):
            anatomy_payload = safe_literal_eval(row["region_attributes"]) or {}
            anatomy_name = str(anatomy_payload.get("Anatomy", "")).strip().lower()
            polygon = parse_via_polygon(row["region_shape_attributes"])
            if anatomy_name in anatomy_map and len(polygon) >= 3:
                polygon_store[row["filename"]][anatomy_map[anatomy_name]].append(polygon)
        segmentation_samples = []
        for filename, polygons in polygon_store.items():
            image_path = image_dir / filename
            if image_path.exists():
                segmentation_samples.append(
                    {
                        "image": str(image_path),
                        "mask_polygons": polygons,
                        "patient_id": _kidney_patient_id(image_path.name),
                    }
                )
        seg_train, seg_val, seg_test = _group_split(segmentation_samples, seed=seed, group_key="patient_id")
        segmentation_task = TaskSpec(
            task_id="kidney/anatomy_segmentation",
            domain="kidney",
            title="Kidney anatomy segmentation",
            task_type=TaskType.SEGMENTATION,
            dataset_names=["OpenKidneyUS"],
            labels=["capsule", "central_echo_complex", "medulla", "cortex"],
            description="Four-region kidney segmentation using the first reviewed annotation set.",
            train_samples=seg_train,
            val_samples=seg_val,
            test_samples=seg_test,
            extras={
                "segmentation_mode": "multiclass",
                "mask_class_names": ["background", "capsule", "central_echo_complex", "medulla", "cortex"],
            },
        )
    else:
        segmentation_task = _build_unavailable_task(
            "kidney/anatomy_segmentation",
            "kidney",
            "Kidney anatomy segmentation",
            TaskType.SEGMENTATION,
            ["OpenKidneyUS"],
            "kidneyUS-main folder was not found.",
        )

    classification_samples: List[Dict] = []
    kidney_class_root = data_root / "Organ_Anomaly" / "organ_classification+anomaly_detection"
    for anomaly_name, label in [("normal", 0), ("abnormal", 1)]:
        folder = kidney_class_root / anomaly_name / "kidney"
        if folder.exists():
            for image_path in _list_images(folder):
                classification_samples.append(
                    {
                        "image": str(image_path),
                        "label": label,
                        "patient_id": f"kidneycls_{anomaly_name}_{image_path.stem}",
                    }
                )
    cls_train, cls_val, cls_test = _group_stratified_split(
        classification_samples,
        seed=seed,
        label_key="label",
        group_key="patient_id",
    )
    classification_task = TaskSpec(
        task_id="kidney/normal_abnormal_classification",
        domain="kidney",
        title="Kidney normal/abnormal classification",
        task_type=TaskType.CLASSIFICATION,
        dataset_names=["Organ Anomaly (kidney subset)"],
        labels=["normal", "abnormal"],
        description="Binary kidney classification built from the kidney slices in Organ Anomaly.",
        train_samples=cls_train,
        val_samples=cls_val,
        test_samples=cls_test,
        notes=["The local OpenKidneyUS copy exposes anatomy labels, not pathology labels, so classification uses the Organ Anomaly kidney subset."],
    )
    return [segmentation_task, classification_task]


def discover_pcos_task(data_root: Path, seed: int) -> TaskSpec:
    train_root = data_root / "PCOS" / "PCOS - Train"
    test_root = data_root / "PCOS" / "PCOS - test"
    if not train_root.exists():
        return _build_unavailable_task(
            "pcos/binary_classification",
            "pcos",
            "PCOS binary classification",
            TaskType.CLASSIFICATION,
            ["PCOS US"],
            "PCOS folders were not found.",
        )

    def row_to_label(row: Dict[str, object]) -> int:
        visible = str(row["Class label (whether polycsytic ovary is visible or not visible)"]).strip().lower()
        return 1 if visible.startswith("visible") else 0

    train_rows = _read_xlsx_rows(train_root / "class_label.xlsx")
    train_pool = []
    for row in train_rows:
        image_name = str(row["imagePath"]).strip()
        image_path = train_root / "images" / image_name
        if image_path.exists():
            train_pool.append(
                {
                    "image": str(image_path),
                    "label": row_to_label(row),
                    "patient_id": _pcos_patient_id(image_name),
                }
            )
    train_samples = list(train_pool)

    test_pool = []
    test_csv = test_root / "class label.csv"
    if test_csv.exists():
        for row in _read_csv_rows(test_csv):
            image_name = row["imagePath"].strip()
            image_path = test_root / "images" / image_name
            if image_path.exists():
                test_pool.append(
                    {
                        "image": str(image_path),
                        "label": row_to_label(row),
                        "patient_id": _pcos_patient_id(image_name),
                    }
                )
    if test_pool:
        val_samples, test_samples = _group_stratified_val_test_split(
            test_pool,
            seed=seed,
            label_key="label",
            group_key="patient_id",
            val_ratio=0.5,
        )
    else:
        _, val_samples, test_samples = _group_stratified_split(
            train_pool,
            seed=seed,
            label_key="label",
            group_key="patient_id",
            val_ratio=0.15,
            test_ratio=0.15,
        )
    return TaskSpec(
        task_id="pcos/binary_classification",
        domain="pcos",
        title="PCOS binary classification",
        task_type=TaskType.CLASSIFICATION,
        dataset_names=["PCOS US"],
        labels=["negative", "positive"],
        description="Binary PCOS classification using the visible/not-visible ovarian label.",
        train_samples=train_samples,
        val_samples=val_samples,
        test_samples=test_samples,
        notes=[
            "Positive is defined here as the spreadsheet's `Visible` label.",
            "Training uses `PCOS - Train`; validation and test are carved patient-level from `PCOS - test` when that folder is available.",
        ],
    )


def discover_carotid_tasks(data_root: Path, seed: int) -> List[TaskSpec]:
    root = data_root / "Carotid" / "Common Carotid Artery Ultrasound Images" / "Common Carotid Artery Ultrasound Images"
    image_dir = root / "US images"
    mask_dir = root / "Expert mask images"
    if not image_dir.exists():
        unavailable = _build_unavailable_task(
            "carotid/lumen_segmentation",
            "carotid",
            "Carotid lumen segmentation",
            TaskType.SEGMENTATION,
            ["Carotid US"],
            "Carotid US image folder was not found.",
        )
        return [unavailable]

    segmentation_samples: List[Dict] = []
    regression_samples: List[Dict] = []
    for image_path in _list_images(image_dir):
        mask_path = mask_dir / image_path.name
        if not mask_path.exists():
            continue
        patient_id = image_path.name.split("_slice_")[0]
        segmentation_samples.append(
            {
                "image": str(image_path),
                "mask_path": str(mask_path),
                "patient_id": patient_id,
            }
        )
        if Image is not None and np is not None:
            mask = np.asarray(Image.open(mask_path).convert("L"), dtype=np.uint8)
            regression_samples.append(
                {
                    "image": str(image_path),
                    "target": carotid_thickness_from_mask(mask),
                    "patient_id": patient_id,
                }
            )
    seg_train, seg_val, seg_test = _group_split(segmentation_samples, seed=seed, group_key="patient_id")
    reg_train, reg_val, reg_test = _group_split(regression_samples, seed=seed, group_key="patient_id")
    measurement_task: TaskSpec
    if reg_train or reg_val or reg_test:
        measurement_task = TaskSpec(
            task_id="carotid/imt_measurement",
            domain="carotid",
            title="Carotid IMT measurement proxy",
            task_type=TaskType.MEASUREMENT,
            dataset_names=["Carotid US"],
            description="Regression over a mask-derived thickness proxy. Units are pixels unless physical spacing is added.",
            train_samples=reg_train,
            val_samples=reg_val,
            test_samples=reg_test,
            extras={"target_name": "thickness_pixels"},
            notes=["The local Carotid copy does not expose calibrated physical spacing, so the target is a pixel-space thickness proxy."],
        )
    else:
        reason = "Install `numpy` and `pillow` to derive measurement targets from the carotid masks."
        measurement_task = _build_unavailable_task(
            "carotid/imt_measurement",
            "carotid",
            "Carotid IMT measurement proxy",
            TaskType.MEASUREMENT,
            ["Carotid US"],
            reason,
        )

    return [
        TaskSpec(
            task_id="carotid/lumen_segmentation",
            domain="carotid",
            title="Carotid lumen segmentation",
            task_type=TaskType.SEGMENTATION,
            dataset_names=["Carotid US"],
            labels=["lumen"],
            description="Binary carotid segmentation from expert mask overlays.",
            train_samples=seg_train,
            val_samples=seg_val,
            test_samples=seg_test,
            extras={"segmentation_mode": "binary"},
        ),
        measurement_task,
    ]


def discover_all_tasks(data_root: Path, seed: int = 42) -> List[TaskSpec]:
    tasks: List[TaskSpec] = []
    tasks.extend(discover_butterfly_tasks(data_root, seed))
    tasks.extend(discover_breast_tasks(data_root, seed))
    tasks.extend(discover_thyroid_tasks(data_root, seed))
    tasks.extend(discover_fetal_tasks(data_root, seed))
    tasks.extend(discover_abdominal_tasks(data_root, seed))
    tasks.extend(discover_liver_tasks(data_root, seed))
    tasks.extend(discover_kidney_tasks(data_root, seed))
    tasks.append(discover_pcos_task(data_root, seed))
    tasks.extend(discover_carotid_tasks(data_root, seed))
    tasks.append(discover_domain_router_task(tasks, seed))
    return tasks


def discover_domain_router_task(tasks: Sequence[TaskSpec], seed: int) -> TaskSpec:
    train_samples: List[Dict] = []
    val_samples: List[Dict] = []
    test_samples: List[Dict] = []
    available_domains: List[str] = []
    missing_domains: List[str] = []
    all_domains = [domain for domain in DOMAIN_ROUTER_DOMAINS if any(task.domain == domain for task in tasks)]
    label_map = {domain: index for index, domain in enumerate(all_domains)}

    for domain in all_domains:
        domain_samples = _aggregate_router_domain_samples(tasks, domain=domain, seed=seed + label_map[domain])
        if len(domain_samples) < 8:
            missing_domains.append(domain)
            continue
        available_domains.append(domain)
        domain_train, domain_val, domain_test = _group_split(
            domain_samples,
            seed=seed + label_map[domain],
            group_key="patient_id",
        )
        for split_target, split_samples in (
            (train_samples, domain_train),
            (val_samples, domain_val),
            (test_samples, domain_test),
        ):
            for sample in split_samples:
                split_target.append(
                    {
                        "image": sample["image"],
                        "label": label_map[domain],
                        "patient_id": sample["patient_id"],
                        "source_task": sample.get("source_task", ""),
                    }
                )

    if len(available_domains) < 2:
        reason = "At least two ready domain datasets are required to build the global domain router."
        if missing_domains:
            reason += f" Missing: {', '.join(sorted(missing_domains))}."
        return _build_unavailable_task(
            "system/domain_classification",
            "system",
            "Global domain routing classifier",
            TaskType.CLASSIFICATION,
            ["PULSE aggregated domains"],
            reason,
        )

    label_names = [domain for domain in all_domains if domain in set(available_domains)]
    compact_label_map = {domain: index for index, domain in enumerate(label_names)}
    for split_samples in (train_samples, val_samples, test_samples):
        for sample in split_samples:
            domain = all_domains[sample["label"]] if sample["label"] < len(all_domains) else None
            if domain is None:
                continue
            sample["label"] = compact_label_map[domain]

    task = TaskSpec(
        task_id="system/domain_classification",
        domain="system",
        title="Global domain routing classifier",
        task_type=TaskType.CLASSIFICATION,
        dataset_names=["PULSE aggregated domains"],
        labels=label_names,
        description="Balanced multi-domain router built from patient-level multi-source image pools across the ultrasound domains.",
        train_samples=train_samples,
        val_samples=val_samples,
        test_samples=test_samples,
        notes=[
            "Router samples are aggregated across all ready tasks in each domain and then re-split at the patient level to avoid leakage.",
            "Mixed-source tasks are remapped to their true anatomy domain before aggregation, for example Butterfly thyroid/carotid frames are not labeled as cardiac.",
            f"Per-domain sample counts are capped at {DOMAIN_ROUTER_MAX_SAMPLES_PER_DOMAIN} to keep router training balanced.",
        ],
        extras={
            "router_task": True,
            "available_domains": sorted(available_domains),
            "balanced_sampling": True,
            "encoder_width": 32,
            "hidden_dim": 224,
            "dropout": 0.15,
        },
    )
    if missing_domains:
        task.notes.append(f"Domains missing from this router build: {', '.join(sorted(missing_domains))}.")
    return task
